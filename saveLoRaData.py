import serial
import os
import openpyxl 
from datetime import datetime

arduino_port = "COM10"   # Arduino serial port (LoRa RX)
baud = 9600 

serial1 = serial.Serial(arduino_port, baud)
print(f'Connected to Arduino port: {arduino_port}')

samples = 100000   # samples to collect
line = 0   # starts at 0 because our header is 0 (not real data)

planilha = openpyxl.load_workbook("medicoesLoRa.xlsx")   
sheet1 = planilha.active   # select sheet to save data  

while line <= samples:
    getData = str(serial1.readline().decode('utf-8','ignore'))   # get data from serial
    temp = getData[0:][:-2]   # cut out '\n'
    data = temp.split("/")   # split information of the serial 
    print(data)

    date_time = datetime.strftime(datetime.now(),'%Y/%m/%d/%H/%M/%S')   # get time at the moment 
    date = date_time.split("/")

    for i, info1 in enumerate (data):
        cellref = sheet1.cell(row = line+1, column = i+1)
        cellref.value = info1   # writes info into separated cell
    for j, info2 in enumerate (date):
        cellref = sheet1.cell(row = line+1, column = j+1+i+1)
        cellref.value = info2   # writes time into following cell
    planilha.save("medicoesLoRa.xlsx")   # saves editing
    line = line+1   # increments

print("Data collection complete!")
